import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
import pandas as pd
import openpyxl
import os
from os import path


window = tk.Tk()
window.title("xlsxViewer")
window.geometry('500x500')


tab_control = ttk.Notebook(window)
tab_control.pack(expand=1, fill='both')
tab_file = ttk.Frame(tab_control)
tab_control.add(tab_file, text='File')
tab_control.pack(expand=1, fill='both')


sheets = []
lbl_sheets_num = []
btn_sheets = []
tabs_sheets = []


def createTabSheet(sheets_it):
    tabs_sheets.insert(sheets_it, ttk.Frame(tab_control))
    tab_control.add(tabs_sheets[sheets_it], text=excel_file.sheetnames[sheets_it])
    cols_it = 0
    rows_it = 0
    flag1 = False
    flag2 = False
    while True:
        while True:
            try:
                tk.Label(tabs_sheets[sheets_it], text=sheets[sheets_it][cols_it][rows_it],
                         font=("Times", 10)).grid(column=cols_it, row=rows_it)
                flag1 = False
            except KeyError:
                if flag1:
                    flag2 = True
                flag1 = True
                rows_it = 0
                break
            rows_it += 1
        if flag2:
            break
        cols_it += 1


def disableSheet():
    ...


excel_file = openpyxl.workbook.workbook.Workbook
def selectFile():
    file = filedialog.askopenfilename(filetypes=(("Table file", "*.xlsx"), ("all files", "*.*")))
    lbl_filename = tk.Label(tab_file, text=path.splitext(path.basename(file))[0], font=("Times", 20)).grid(column=0, row=1)
    lbl_sheets = tk.Label(tab_file, text="Sheets:", font=("Times", 20)).grid(column=0, row=2)
    sheets_it = 0
    global excel_file
    excel_file = openpyxl.load_workbook(file)
    while True:
        try:
            sheets.insert(sheets_it, pd.read_excel(file, sheet_name=sheets_it, header=None))
        except ValueError:
            break
        lbl_sheets_num.insert(sheets_it, tk.Label(tab_file, text=excel_file.sheetnames[sheets_it],
                                                font=("Times", 15)).grid(column=0, row=3+sheets_it))
        btn_sheets.insert(sheets_it, tk.Button(tab_file, text="Hide", font=("Times", 15),
                                               command=disableSheet).grid(column=1, row=3+sheets_it))
        createTabSheet(sheets_it)
        sheets_it += 1


btn_select_file = tk.Button(tab_file, text="Select a file", command=selectFile, font=("Times", 30)).grid(column=0, row=0)


window.mainloop()